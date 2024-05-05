import openpyxl


class TestData:
    test_data = [{"username": "superadmintravelmate@getairmail.com", "password": "Atlas@123"},
                 {"username": "hrsdonly@mailsac.com", "password": "Atlas@123"}]
    test_data_superAdmin = [{"username": "superadmintravelmate@getairmail.com", "password": "Atlas@123"}]

    @staticmethod
    def getExcelData(testcase):
        book = openpyxl.load_workbook("C://Users//ayyappan.g//PycharmProjects//Atlas_Pyton_Framework//testData.xlsx")
        sheet = book.active
        Dict = {}
        for m in range(1, sheet.max_row + 1):
            if sheet.cell(row=m, column=1).value == testcase:
                for n in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=n).value] = sheet.cell(row=m, column=n).value
        return [Dict]

    @staticmethod
    def getExcelDataTrail():
        book = openpyxl.load_workbook("C://Users//ayyappan.g//PycharmProjects//Atlas_Pyton_Framework//testData.xlsx")
        sheet = book.active

        Listed = []
        for m in range(2, sheet.max_row + 1):
            Dict = {}
            for n in range(2, sheet.max_column + 1):
                Dict[sheet.cell(row=1, column=n).value] = sheet.cell(row=m, column=n).value
            Listed.append(Dict)
        return Listed
