import openpyxl

class Dummy:
    def readexcel(self):
        book = openpyxl.load_workbook("C://Users//ayyappan.g//PycharmProjects//Atlas_Pyton_Framework//dummyExcel.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=1)
        cell2 = sheet.cell(row=1, column=2)
        cell.value = "Sangar"
        cell2.value = "kumaresan"
        book.save("C://Users//ayyappan.g//PycharmProjects//Atlas_Pyton_Framework//dummyExcel.xlsx")
        print(cell.value)
        print(cell2.value)

obj=Dummy()
obj.readexcel()
